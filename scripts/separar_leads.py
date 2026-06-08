"""
Separa os leads do Instagram em abas Excel por curso e por mês/data.
Uso: python3 scripts/separar_leads.py
"""
import csv
import os
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    os.system("pip3 install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_FILE = os.path.join(BASE_DIR, "Cadastros_no_Instagram.csv")
OUTPUT_FILE = os.path.join(BASE_DIR, "Leads_Separados.xlsx")

COLUNAS = ["Form Name", "Qual o curso que você tem interesse?", "Full name", "Email", "Phone number", "Created Time", "Campaign Name", "Stage Name"]
COLUNAS_PT = ["Campanha (Form)", "Curso de Interesse", "Nome Completo", "Email", "Telefone", "Data de Cadastro", "Nome da Campanha", "Etapa"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}


def ler_csv():
    rows = []
    with open(INPUT_FILE, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            data_raw = r.get("Created Time", "")
            try:
                dt = datetime.fromisoformat(data_raw)
                r["_dt"] = dt
                r["_mes_ano"] = f"{MESES_PT[dt.month]} {dt.year}"
                r["_mes_sort"] = dt.strftime("%Y-%m")
                r["_data_br"] = dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                r["_dt"] = None
                r["_mes_ano"] = "Sem data"
                r["_mes_sort"] = "9999-99"
                r["_data_br"] = data_raw
            rows.append(r)
    return rows


def formatar_header(ws, colunas):
    ws.append(colunas)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30


def adicionar_linha(ws, row_idx, valores):
    ws.append(valores)
    fill = ALT_FILL if row_idx % 2 == 0 else None
    for cell in ws[ws.max_row]:
        if fill:
            cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")


def ajustar_colunas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def linha_para_valores(row):
    return [
        row.get("Form Name", ""),
        row.get("Qual o curso que você tem interesse?", ""),
        row.get("Full name", ""),
        row.get("Email", ""),
        row.get("Phone number", ""),
        row.get("_data_br", ""),
        row.get("Campaign Name", ""),
        row.get("Stage Name", ""),
    ]


def criar_aba_resumo(wb, rows):
    ws = wb.active
    ws.title = "Resumo"

    ws.append([])
    ws.merge_cells("A1:H1")
    titulo = ws["A1"]
    titulo.value = "LEADS INSTAGRAM — RESUMO GERAL"
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor="1F4E79")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Por curso
    ws.append([])
    ws["A2"] = "Por Curso"
    ws["A2"].font = Font(bold=True, size=12, color="1F4E79")

    ws.append(["Curso", "Total de Leads"])
    for cell in ws[3]:
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = BORDER

    por_curso = defaultdict(int)
    for r in rows:
        curso = r.get("Qual o curso que você tem interesse?", "") or "(sem curso)"
        por_curso[curso] += 1

    for i, (curso, total) in enumerate(sorted(por_curso.items(), key=lambda x: -x[1]), start=4):
        ws.append([curso, total])
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in ws[i]:
            if fill:
                cell.fill = fill
            cell.border = BORDER

    # Espaço
    linha_atual = ws.max_row + 2
    ws.cell(row=linha_atual, column=1, value="Por Mês")
    ws.cell(row=linha_atual, column=1).font = Font(bold=True, size=12, color="1F4E79")

    linha_atual += 1
    ws.cell(row=linha_atual, column=1, value="Mês")
    ws.cell(row=linha_atual, column=2, value="Total de Leads")
    for col in [1, 2]:
        cell = ws.cell(row=linha_atual, column=col)
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = BORDER

    por_mes = defaultdict(lambda: {"total": 0, "sort": ""})
    for r in rows:
        mes = r["_mes_ano"]
        por_mes[mes]["total"] += 1
        por_mes[mes]["sort"] = r["_mes_sort"]

    for i, (mes, info) in enumerate(sorted(por_mes.items(), key=lambda x: x[1]["sort"]), start=linha_atual + 1):
        ws.cell(row=i, column=1, value=mes)
        ws.cell(row=i, column=2, value=info["total"])
        fill = ALT_FILL if i % 2 == 0 else None
        for col in [1, 2]:
            cell = ws.cell(row=i, column=col)
            if fill:
                cell.fill = fill
            cell.border = BORDER

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18


def criar_aba_todos(wb, rows):
    ws = wb.create_sheet("Todos os Leads")
    formatar_header(ws, COLUNAS_PT)
    for i, row in enumerate(rows, start=1):
        adicionar_linha(ws, i, linha_para_valores(row))
    ajustar_colunas(ws)
    ws.freeze_panes = "A2"


def criar_abas_por_curso(wb, rows):
    por_curso = defaultdict(list)
    for r in rows:
        curso = r.get("Qual o curso que você tem interesse?", "").strip() or "Sem Curso"
        por_curso[curso].append(r)

    for curso, leads in sorted(por_curso.items(), key=lambda x: -len(x[1])):
        nome_aba = curso[:31]  # Excel limita a 31 chars
        ws = wb.create_sheet(nome_aba)
        formatar_header(ws, COLUNAS_PT)
        leads_ord = sorted(leads, key=lambda r: r["_mes_sort"])
        for i, row in enumerate(leads_ord, start=1):
            adicionar_linha(ws, i, linha_para_valores(row))
        ajustar_colunas(ws)
        ws.freeze_panes = "A2"


def criar_abas_por_mes(wb, rows):
    por_mes = defaultdict(list)
    for r in rows:
        por_mes[(r["_mes_sort"], r["_mes_ano"])].append(r)

    for (sort_key, mes_nome), leads in sorted(por_mes.items(), key=lambda x: x[0][0]):
        nome_aba = mes_nome[:31]
        ws = wb.create_sheet(nome_aba)
        formatar_header(ws, COLUNAS_PT)
        leads_ord = sorted(leads, key=lambda r: r["_dt"] or datetime.min)
        for i, row in enumerate(leads_ord, start=1):
            adicionar_linha(ws, i, linha_para_valores(row))
        ajustar_colunas(ws)
        ws.freeze_panes = "A2"


def main():
    print(f"Lendo: {INPUT_FILE}")
    rows = ler_csv()
    print(f"Total de leads: {len(rows)}")

    wb = openpyxl.Workbook()

    print("Criando aba Resumo...")
    criar_aba_resumo(wb, rows)

    print("Criando aba Todos os Leads...")
    criar_aba_todos(wb, rows)

    print("Criando abas por curso...")
    criar_abas_por_curso(wb, rows)

    print("Criando abas por mês...")
    criar_abas_por_mes(wb, rows)

    wb.save(OUTPUT_FILE)
    print(f"\nArquivo gerado: {OUTPUT_FILE}")

    cursos = set(r.get("Qual o curso que você tem interesse?", "") or "Sem Curso" for r in rows)
    meses = set(r["_mes_ano"] for r in rows)
    print(f"  Abas de cursos: {len(cursos)}")
    print(f"  Abas de meses: {len(meses)}")
    print(f"  Total de abas: {2 + len(cursos) + len(meses)} (Resumo + Todos + cursos + meses)")


if __name__ == "__main__":
    main()
